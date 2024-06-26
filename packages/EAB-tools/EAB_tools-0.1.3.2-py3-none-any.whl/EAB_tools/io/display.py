"""Methods to display and save DataFrames, plots."""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from typing import Any
import warnings

from IPython.display import display

with warnings.catch_warnings():
    warnings.simplefilter("ignore", DeprecationWarning)
    import dataframe_image as dfi  # noqa

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from pandas.io.formats.style import (
    Styler,
    Subset,
)

from EAB_tools.io.filenames import (
    PathLike,
    sanitize_filename,
    sanitize_xl_sheetname,
)
from EAB_tools.util.hashing import (
    hash_df,
    hash_mpl_fig,
)

# Copied from Excel's conditional formatting Red-Yellow-Green built-in colormap.
_xl_RYG_colors = ["#F8696B", "#FFEB84", "#63BE7B"]
xl_RYG_cmap = mpl.colors.LinearSegmentedColormap.from_list(
    "xl_RYG_cmap", _xl_RYG_colors
)
xl_GYR_cmap = xl_RYG_cmap.reversed()


def _to_excel(
    df: pd.DataFrame,
    styler: Styler,
    excel_output: Path,
    sheet_name: PathLike | None,
    percentage_format_subset: Subset | str | None,
    thousands_format_subset: Subset | str | None,
    bar_subset: Subset | str | None,
    date_format: str,
    percentage_format_precision: int,
    bar_vmin: float | None,
    bar_vmax: float | None,
) -> None:
    """Export DataFrame to Excel. Used as a helper function to display_and_save_df."""
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl is required for Excel functionality") from e

    # Determine ExcelWriter params based on if the file exists or not
    mode, if_sheet_exists = ("a", "replace") if excel_output.exists() else ("w", None)

    # Determine an Excel sheet name:
    if sheet_name is None:
        sheet_name = excel_output.name.replace(".png", "").replace(".xlsx", "")
    sheet_name = Path(sheet_name)
    sheet_name = sheet_name.name.replace(".df.png", "").replace(".png", "")
    sheet_name = sanitize_xl_sheetname(sheet_name)

    # Excel does NOT support datetimes with timezones
    for col in df.select_dtypes(["datetime", "datetimetz"]).columns:
        df[col] = df[col].dt.tz_localize(None)
        styler.data[col] = df[col]
        styler = styler.format(f"{{:{date_format}}}", subset=col)

    # Export to Excel:
    with pd.ExcelWriter(
        excel_output, engine="openpyxl", mode=mode, if_sheet_exists=if_sheet_exists
    ) as wb:
        print(
            f"Exporting to Excel as '{excel_output.resolve().parent}\\"
            f"[{excel_output.name}]{sheet_name}'",
            end=" ... ",
        )
        styler.to_excel(wb, sheet_name=sheet_name, engine="openpyxl")

        if (
            percentage_format_subset is not None
            or thousands_format_subset is not None
            or bar_subset is not None
        ):
            # Number formatting doesn't seem to carry over to
            # Excel automatically with pandas. Since percentages, thousands, etc.
            # are so widespread, we are using openpyxl to convert the number
            # formats.

            # Additionally, we are using openpyxl to add data bar conditional
            # formatting for bar_subset.
            if isinstance(percentage_format_subset, str):
                percentage_format_subset = [percentage_format_subset]
            if isinstance(thousands_format_subset, str):
                thousands_format_subset = [thousands_format_subset]
            if isinstance(bar_subset, str):
                bar_subset = [bar_subset]
            len_index = df.index.nlevels

            # Get the worksheet
            sheet_name = next(
                sheet
                for sheet in wb.book.sheetnames
                if sheet.casefold() == sheet_name.casefold()
            )
            ws: openpyxl.workbook.workbook.Worksheet = wb.book[sheet_name]

            # Determine the 0-based pd indices for number formatting
            pcnt_cols = (
                [df.columns.get_loc(col) for col in percentage_format_subset]
                if percentage_format_subset is not None
                else []
            )
            try:
                # Check if we have a list containing a list
                iter(pcnt_cols[0])
                pcnt_cols = pcnt_cols[0]
            except (TypeError, IndexError):
                # We just have a normal list
                pass
            tsnd_cols = (
                [df.columns.get_loc(col) for col in thousands_format_subset]
                if thousands_format_subset is not None
                else []
            )
            try:
                if isinstance(bar_subset, (list, tuple)) and len(bar_subset) > 1:
                    # If `bar_subset` is a tuple, then it's
                    # (rows, cols)
                    bar_subset = bar_subset[1]
                bar_cols = (
                    [df.columns.get_loc(col) for col in bar_subset]
                    if bar_subset is not None
                    else []
                )
            except TypeError:
                # `bar_subset`` contains more than just columns.
                # For now, we must abort trying to export data bar conditional
                # formatting.
                bar_cols = []

            # Iterate through the columns, applying styles to all
            # cells in a column
            for col_num, col in enumerate(ws.iter_cols()):
                # Apply percentage, thousands number formats
                if np.any(col_num - len_index in pcnt_cols):
                    for cell in col:
                        code = (
                            f"0{'.' * (percentage_format_precision > 0)}"
                            f"{'0' * percentage_format_precision}%"
                        )
                        cell.number_format = code
                if col_num - len_index in tsnd_cols:
                    for cell in col:
                        cell.number_format = "#,##0"

                # Add bar conditional formatting, using the style's vmin and vmax,
                # if given
                rule = openpyxl.formatting.rule.DataBarRule(
                    start_type="num" if bar_vmin else "min",
                    start_value=bar_vmin,
                    end_type="num" if bar_vmax else "max",
                    end_value=bar_vmax,
                    color="638EC6",
                    showValue=True,
                    minLength=None,
                    maxLength=None,
                )
                for bar_col in bar_cols:
                    letter = openpyxl.utils.cell.get_column_letter(
                        bar_col + len_index + 1
                    )
                    end_num = len(df) + df.columns.nlevels
                    xl_range = f"{letter}1:{letter}{end_num}"

                    # I could not even find the function
                    # ws.conditional_formatting.add() in the openpyxl docs. Thank
                    # god for https://stackoverflow.com/a/32454012.
                    ws.conditional_formatting.add(xl_range, rule)


def display_and_save_df(
    df: pd.DataFrame | pd.Series | Styler,
    caption: str | None = None,
    filename: PathLike | None = None,
    large_title: bool = True,
    large_col_names: bool = True,
    cell_borders: bool = True,
    highlight_total_row: bool = False,
    border_width: str = "1px",
    thousands_format_subset: Subset | str | None = "auto",
    date_format_subset: Subset | str | None = "auto",
    date_format: str = "%#m/%#d/%Y",
    percentage_format_subset: Subset | str | None = "auto",
    percentage_format_precision: int = 1,
    float_format_subset: Subset | str | None = "auto",
    float_format_precision: int = 1,
    hide_index: bool = False,
    convert_dtypes: bool = True,
    ryg_bg_subset: Subset | str | None = None,
    ryg_bg_vmin: float | None = None,
    ryg_bg_vmax: float | None = None,
    gyr_bg_subset: Subset | str | None = None,
    gyr_bg_vmin: float | None = None,
    gyr_bg_vmax: float | None = None,
    bar_subset: Subset | str | None = None,
    bar_vmin: float | None = None,
    bar_vmax: float | None = None,
    format_kwargs: Sequence[dict[str, Any]] | None = None,
    background_gradient_kwargs: Sequence[dict[str, Any]] | None = None,
    bar_kwargs: Sequence[dict[str, Any]] | None = None,
    save_excel: bool = False,
    excel_path: PathLike = "output.xlsx",
    save_image: bool = False,
    min_width: str = "10em",
    max_width: str = "25em",
) -> Styler:
    """
    Display and save a pandas DataFrame or Styler object.

    This functions provides lots of pre-built styling for a pandas DataFrame, with
    highly configurable additional options. Optionally, it can save the DataFrame as
    a png image or export to Excel.

    Parameters
    ----------
    df : pandas.DataFrame, pandas.Series or pd.io.formats.style.Styler
        The pandas object to display and save.
    caption : str, optional
        The caption to be used on the HTML table. The value of caption will also be
        used for the filename and Excel sheetname if filename is not specified and
        either save_excel or save_image is True.
    filename : str, optional
        When saving to Excel or image, the filename to use. When the filename is not
        specified, the value of caption will be used.
    large_title : bool, default True
        Whether to make the title large on the HTML table. Currently, it sets the
        title fontsize to "225%" and cannot be adjusted at this time.
    large_col_names : bool, default True
        Whether to make column headers large. Currently, it sets the column headers
        fontsize to "100%", the "boarder-style" to "solid", and the alignment to
        "center".
    cell_borders : bool, default True
        Whether to draw solid cell borders. The border width is determined by
        `boarder_width`.
    highlight_total_row : bool, default False
        Whether to highlight the final row on the table to give it the appearance of a
        grand totals row. Currently, sets the font to "bold" and the fontsize to "110%".
    border_width : str, default "1px"
        If cell_borders is True, then use this value for the border width. Must be a
        valid HTML border-width.

        Note that if cell_borders is False, this value is completely ignored.
    thousands_format_subset : slice, sequence or str, optional, default "auto"
        The subset to format with a thousand's seperator. Either a slice, a single
        column name, or a sequence of column names. If left to its default value of
        "auto", then all columns of dtype `int` are selected.
    date_format_subset : slice, sequence or str, optional, default "auto"
        The subset to format with a call to `strftime`. Either a slice, a single column
        name, or a sequence of column names. If left to its default value of "auto",
        then all columns of dtype "datetime" are selected.
    date_format : str, default "%#m/%#d/%Y"
        If date_format_subset is not None, the value to use in calls to `strftime` for
        the data_format_subset. Must be a valid strftime-code.
    percentage_format_subset : slice, sequence or str, optional, default "auto"
        The subset to format as a percentage. Either a slice, a single column name, or a
        sequence of column names. If left to its default value of "auto", it will
        attempt to automatically select all columns containing a literal % or the
        literal word "percent". This automatic selection is not fool-proof and may not
        work for all types of MultiIndex objects.
    percentage_format_precision : int, default 1
        if percentage_format_subset is not None, the number of decimal places to show in
        each column that is percentage formatted.
    float_format_subset : slice, sequence or str, optional, default "auto"
        The subset to format as a float. These columns will obey the
        float_format_precision provided. Either a slice, a single column name, or a
        sequence of column names. If left to its default value of "auto", it selects all
        columns with dtype float that are *not* in the percentage_format_subset.
    float_format_precision : int, default 1
        If float_format_subset is not None, the precision to use when displaying
        values from the subset.
    hide_index : bool, default False
        Whether to hide the index in the final HTML table. This is useful when the
        index has no meaningful interpretation.
    convert_dtypes : bool, default True
        Attempt to convert columns to their optimal dtypes using
        `pandas.DataFrame.convert_dtypes`. This can be useful for ensuring that
        automatic subsets work correctly, but can cause errors in cases where pandas
        incorrectly infers the dtypes.
    ryg_bg_subset : slice, subset or str, optional
        The subset to format with "conditional formatting", going from red -> yellow ->
        green in ascending order. It is meant to resemble Excel's build-in RYG
        conditional formatting. Either a slice, a single column name, or a sequence of
        column names.
    ryg_bg_vmin : float, optional
        If ryg_bg_subset is not None, the minimum data value that corresponds to the
        minimum red value.
    ryg_bg_vmax : float, optional
        If ryg_bg_subset is not None, the maximum data value that corresponds to the
        maximum green value.
    gyr_bg_subset : slice, subset or str, optional
        The subset to format with "conditional formatting", going from green -> yellow
        -> red in ascending order. It is meant to resemble Excel's build-in GYR
         conditional formatting. Either a slice, a single column name, or a sequence of
         column names.
    gyr_bg_vmin : float, optional
        If gyr_bg_subset is not None, the minimum data value that corresponds to the
        minimum green color.
    gyr_bg_vmax : float, optional
        If gyr_bg_subset is not None, the maximum data value that corresponds to the
        maximum red color.
    bar_subset : slice, subset or str, optional
        The subset to format with bar charts in the cell backgrounds. Either a slice, a
        single column name, or a sequence of column names.
    bar_vmin : float, optional
        If bar_subset is not None, the minimum data value defining the left-hand limit
        of the bar drawing range.
    bar_vmax : float, optional
        If bar_subset is not None, the maximum data value defining the right-hand limit
        of the bar drawing range.
    format_kwargs : dict or sequence of dicts, optional
        When the above parameters are not sufficient, each of the dicts in this list
        will be passed to Styler.format in order as keyword arguments.
    background_gradient_kwargs : dict or sequence of dicts, optional
        When the above parameters are not sufficient, each of the dicts in this list
        will be passed to `Styler.background_gradient` in order as keyword arguments.
    bar_kwargs : dict or sequence of dicts, optional
        When the above parameters are not sufficient, each of the dicts in this list
        will be passed to `Styler.bar` in order as keyword arguments.
    save_excel : bool, default False
        Whether to save the output to an Excel sheet. If True, it will create an Excel
        workbook named "output.xlsx" in the same directory as filename. The caption is
        used for the sheetname, and may be corerced to a valid Excel sheetname.

        Subsequent calls to display_and_save_df in the same directory will *append* to
        the workbook "output.xlsx" if it already exists. If a sheet with the same name
        already exists, it will be overwritten.
    save_image : bool, default False
        Whether to save the output to a png image using the package dataframe_image. If
        True, it will create a png image based on the filename or filepath.
    min_width : str, default "10em"
        The minimum width to use when drawing the HTML table. Must be a valid HTML
        measurement.
    max_width : str, default "25em"
        The maximum width to use when drawing the HTML table. Must be a valid HTML
        measurement.

    Returns
    -------
    pandas.io.formats.style.Styler
        The generated Styler object.

    Examples
    --------
    >>> df = pd.DataFrame({'Letter': list('ABC'), 'Number': range(1, 4)})
    >>> df
      Letter  Number
    0      A       1
    1      B       2
    2      C       3
    >>> display_and_save_df(df)
    <pandas.io.formats.style.Styler object at 0x...>
    <pandas.io.formats.style.Styler object at 0x...>
    >>> styler = df.style.format(na_rep="Omg! Empty!")
    >>> display_and_save_df(styler, hide_index=True)
    <pandas.io.formats.style.Styler object at 0x...>
    <pandas.io.formats.style.Styler object at 0x...>
    """
    if hasattr(df, "copy"):
        df = df.copy(deep=True)
    if format_kwargs is None:
        format_kwargs = []
    if background_gradient_kwargs is None:
        background_gradient_kwargs = []
    if bar_kwargs is None:
        bar_kwargs = []

    # Convert pd.Series to Frame if needed
    if isinstance(df, pd.Series):
        df = df.to_frame()

    # Determine the DataFrame and the Styler
    if isinstance(df, Styler):
        df, styler = df.data, df
    else:
        styler = df.style

    # Convert dtypes if requested
    if convert_dtypes:
        df = df.convert_dtypes()

    # Define some styles.
    # These do NOT export to Excel!
    LARGE_TITLE = {
        "selector": "caption",
        "props": [("font-size", "225%"), ("text-align", "center")],
    }

    LARGE_COL_NAMES = {
        "selector": "th",
        "props": [
            ("font-size", "110%"),
            ("border-style", "solid"),
            ("text-align", "center"),
        ],
    }

    CELL_BORDERS = {
        "selector": "th,td",
        "props": [("border-style", "solid"), ("border-width", border_width)],
    }

    HIGHLIGHT_TOTAL = {
        "selector": "tr:last-child",
        "props": [("font-weight", "bold"), ("font-size", "110%")],
    }

    MIN_MAX_WIDTH = {
        "selector": "th",
        "props": [("min-width", min_width), ("max-width", max_width)],
    }
    # Mind width defaults to "10em" based on hardcoded value in
    # Styler.bar().
    # https://github.com/pandas-dev/pandas/blob/9222cb0c/pandas/io/formats/style.py#L3097

    # Enforce min and max width
    styler = styler.set_table_styles([MIN_MAX_WIDTH], overwrite=False)

    # Apply the caption if it is not None
    if caption is not None:
        styler.set_caption(caption)

    # Apply optional styles
    if large_title:
        styler = styler.set_table_styles([LARGE_TITLE], overwrite=False)
    if large_col_names:
        styler = styler.set_table_styles([LARGE_COL_NAMES], overwrite=False)
    if cell_borders:
        styler = styler.set_table_styles([CELL_BORDERS], overwrite=False)
    if highlight_total_row:
        styler = styler.set_table_styles([HIGHLIGHT_TOTAL], overwrite=False)

    # Find auto percentage format columns
    if isinstance(percentage_format_subset, str) and percentage_format_subset == "auto":
        try:
            flattened_cols = df.columns.to_flat_index().astype("string")
            # Grab the cols that contain percent signs
            regex = "%|percent"  # Either literal % or the word "percent"
            percentage_format_subset_mask = flattened_cols.str.contains(
                regex, case=False
            )
            # But don't accept cols that are a string dtype
            percentage_format_subset_mask &= [
                not pd.api.types.is_string_dtype(df[col]) for col in df
            ]
            percentage_format_subset = df.columns[percentage_format_subset_mask]
        except AttributeError as e:
            # Can only use .str accessor with Index, not MultiIndex
            warnings.warn(str(e), stacklevel=2)
            percentage_format_subset = []
    # Apply the percentage format
    if percentage_format_subset is not None:
        formatter = f"{{:.{percentage_format_precision}%}}"
        styler = styler.format(formatter=formatter, subset=percentage_format_subset)

    # Apply thousands seperator
    if thousands_format_subset is not None:
        if thousands_format_subset == "auto":
            thousands_format_subset = df.select_dtypes(int).columns
        styler = styler.format("{:,}", subset=thousands_format_subset)

    # Apply floating point precision
    if float_format_subset is not None:
        if float_format_subset == "auto":
            float_format_subset = df.select_dtypes(float).columns.drop(
                percentage_format_subset, errors="ignore"
            )
        styler = styler.format(
            f"{{:.{float_format_precision}f}}", subset=float_format_subset
        )

    # Apply date formatting
    if date_format_subset is not None:
        if date_format_subset == "auto":
            date_format_subset = df.select_dtypes("datetime").columns
        styler = styler.format(f"{{:{date_format}}}", subset=date_format_subset)

    # Hide axes
    if hide_index:
        styler = styler.hide(axis="index")

    # Apply RYG or GYR conditional formatting
    if ryg_bg_subset is not None:
        styler = styler.background_gradient(
            xl_RYG_cmap,
            subset=ryg_bg_subset,
            text_color_threshold=0,
            vmin=ryg_bg_vmin,
            vmax=ryg_bg_vmax,
        )
    if gyr_bg_subset is not None:
        styler = styler.background_gradient(
            xl_GYR_cmap,
            subset=gyr_bg_subset,
            text_color_threshold=0,
            vmin=gyr_bg_vmin,
            vmax=gyr_bg_vmax,
        )

    # Apply the histogram bar conditional formatting
    if bar_subset is not None:
        styler = styler.bar(
            subset=bar_subset, color="#638ec6", vmin=bar_vmin, vmax=bar_vmax, width=90
        )

    # Accept a list of kwargs to push through various formatter functions
    kwargs = {
        "format": format_kwargs,
        "background_gradient": background_gradient_kwargs,
        "bar": bar_kwargs,
    }
    # Sometimes a single dict is passed instead of a list of dicts
    kwargs = {
        k: [kwarg] if isinstance(kwarg, dict) else kwarg for k, kwarg in kwargs.items()
    }
    for format_kwarg in kwargs.pop("format"):
        styler = styler.format(**format_kwarg)
    for background_gradient_kwarg in kwargs.pop("background_gradient"):
        # By default, text_color_threshold should be 0. Everything in black text.
        text_color_threshold = background_gradient_kwarg.get(
            "text_color_threshold", 0.0
        )
        styler = styler.background_gradient(
            text_color_threshold=text_color_threshold, **background_gradient_kwarg
        )
    for bar_kwarg in kwargs.pop("bar"):
        color = bar_kwarg.pop("color", "#638ec6")  # Default color from Excel
        styler = styler.bar(color=color, **bar_kwarg)

    # Determine a suitable filename/Excel sheet name
    if filename is None and styler.caption is not None:
        # If no filename is given, use the caption
        filename = f"{styler.caption}.df.png"
        filename = sanitize_filename(filename)
    filename = Path(filename if filename else f"{hash_df(df, styler)}.df.png")

    # Save the Styler as a png
    if save_image:
        filename.parent.mkdir(exist_ok=True, parents=True)

        print(f"Saving as '{filename.resolve()}'", end=" ... ")
        styler.export_png(str(filename), fontsize=16, max_rows=200, max_cols=200)

    # Save the Styler to Excel sheet
    if save_excel:
        _to_excel(
            df=df,
            styler=styler,
            excel_output=Path(excel_path),
            sheet_name=filename,
            percentage_format_subset=percentage_format_subset,
            thousands_format_subset=thousands_format_subset,
            bar_subset=bar_subset,
            date_format=date_format,
            percentage_format_precision=percentage_format_precision,
            bar_vmin=bar_vmin,
            bar_vmax=bar_vmax,
        )

    # Finally, display with styler with IPython
    display(styler)

    return styler


def display_and_save_fig(
    fig: plt.Figure | plt.Axes,
    filename: str | None = None,
    save_image: bool = False,
) -> None:
    """
    Display and save a `matplotlib` figure.

    Displays and saves and `matplotlib` `Figure` or `Axes` with the default `matplotlib`
    backend and optionally saves the image to disk as a png.

    Parameters
    ----------
    fig : `plt.Figure` or `plt.Axes`
        The `matplotlib` object to display and save.
    filename : str, optional
        The filename to use if saving as an image.
    save_image : bool, default False
        If True, save the `matplotlib` plot to disk.

    See Also
    --------
    display_and_save_df : Display and save a pandas DataFrame or Styler object.

    Examples
    --------
    >>> X = np.linspace(0, 2 * np.pi, 10 ** 6)
    >>> y = np.tan(X)
    >>> fig, ax = plt.subplots()
    >>> ax.plot(X, y, "b-")
    [<matplotlib.lines.Line2D object at 0x...>]
    >>> display_and_save_fig(fig, "output image.png", save_image=True)
    Saving as output image.png ...
    >>> display_and_save_fig(ax, "output another image", save_image=True)
    Saving as output another image.png ...
    """
    plt.show(block=False)

    if not save_image:
        # If we're not saving the image, then this is the end of the function.
        return

    # If they passed an `Axes` object, get the `Figure`
    figure = fig.get_figure() if isinstance(fig, plt.Axes) else fig
    assert figure is not None  # for `mypy`

    # Attempt to infer the filename if it is None
    if filename is None:
        if figure._suptitle is not None:  # type: ignore[attr-defined]
            filename = figure._suptitle.get_text()  # type: ignore[attr-defined]
        elif figure.axes[0].title.get_text() != "":
            filename = figure.axes[0].title.get_text()
        else:
            filename = hash_mpl_fig(figure)
    filename = sanitize_filename(filename)

    # Needed to make `mypy` happy
    filename = str(filename)
    if filename[-4:] != ".png":
        filename = filename + ".png"

    filepath = Path(filename)

    print(f"Saving as {filepath} ... ", end="")
    figure.savefig(filepath)
